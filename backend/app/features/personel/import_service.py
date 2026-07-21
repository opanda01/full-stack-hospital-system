"""Personel toplu import — dosya parse, satır işleme, progress."""

from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlmodel import Session, select

from app.core.audit import denetim_kaydi_yaz
from app.core.base_model import utc_now
from app.core.enums import ImportDurum, Rol
from app.core.notifications import get_bildirim
from app.core.security import hash_password
from app.features.auth.models import PersonelImportIsi
from app.features.auth.service import gecici_sifre_uret
from app.features.departmanlar.models import Departman
from app.features.doktorlar.models import Doktor
from app.features.kullanicilar.models import Kullanici
from app.features.personel.models import Personel

# Türkçe / İngilizce kolon alias'ları
_COLUMN_MAP = {
    "ad": "ad",
    "soyad": "soyad",
    "tc": "tc_kimlik_no",
    "tc kimlik no": "tc_kimlik_no",
    "tc_kimlik_no": "tc_kimlik_no",
    "sicil": "sicil_no",
    "sicil no": "sicil_no",
    "sicil_no": "sicil_no",
    "rol": "rol",
    "departman": "departman",
    "telefon": "telefon",
    "email": "email",
    "e-posta": "email",
    "kullanici_adi": "kullanici_adi",
    "kullanıcı adı": "kullanici_adi",
    "uzmanlik": "uzmanlik_alani",
    "uzmanlik_alani": "uzmanlik_alani",
    "uzmanlık alanı": "uzmanlik_alani",
    "diploma_no": "diploma_no",
    "diploma no": "diploma_no",
}

PERSONEL_ROLLERI = frozenset(
    {
        Rol.ADMIN,
        Rol.BASHEKIM,
        Rol.MUDUR,
        Rol.DOKTOR,
        Rol.HEMSIRE,
        Rol.EBE,
        Rol.LABORANT,
        Rol.TEMIZLIK_PERSONELI,
        Rol.GUVENLIK,
        Rol.IDARI_PERSONEL,
    }
)


def _normalize_header(raw: str) -> str | None:
    key = raw.strip().lower()
    return _COLUMN_MAP.get(key)


def parse_import_file(filename: str, content: bytes) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Desteklenen formatlar: CSV, XLSX",
    )


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV başlık satırı eksik")
    rows: list[dict[str, str]] = []
    for raw in reader:
        mapped: dict[str, str] = {}
        for k, v in raw.items():
            if k is None:
                continue
            nk = _normalize_header(k)
            if nk and v is not None and str(v).strip():
                mapped[nk] = str(v).strip()
        if mapped:
            rows.append(mapped)
    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="XLSX desteği için openpyxl kurulu olmalı",
        ) from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel dosyası boş") from None
    headers: list[str | None] = []
    for h in header_row:
        if h is None:
            headers.append(None)
        else:
            headers.append(_normalize_header(str(h)))
    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        mapped: dict[str, str] = {}
        for idx, cell in enumerate(raw):
            if idx >= len(headers) or headers[idx] is None:
                continue
            if cell is None or str(cell).strip() == "":
                continue
            mapped[headers[idx]] = str(cell).strip()  # type: ignore[index]
        if mapped:
            rows.append(mapped)
    return rows


def create_import_isi(
    session: Session,
    *,
    actor_id: int,
    rows: list[dict[str, str]],
    ip_adresi: str | None = None,
) -> PersonelImportIsi:
    isi = PersonelImportIsi(
        actor_id=actor_id,
        durum=ImportDurum.BEKLEMEDE,
        toplam=len(rows),
        basarili=0,
        basarisiz=0,
        hata_detay=[],
    )
    session.add(isi)
    session.commit()
    session.refresh(isi)
    denetim_kaydi_yaz(
        session,
        aksiyon="PERSONEL_IMPORT_BASLAT",
        actor_id=actor_id,
        kaynak="personel_import",
        kaynak_id=isi.id,
        ip_adresi=ip_adresi,
        detay={"toplam": len(rows)},
    )
    return isi


def get_import_isi(session: Session, isi_id: int) -> PersonelImportIsi:
    isi = session.get(PersonelImportIsi, isi_id)
    if isi is None:
        raise HTTPException(status_code=404, detail="Import işi bulunamadı")
    return isi


def _departman_bul(session: Session, ad: str | None) -> int | None:
    if not ad:
        return None
    dep = session.exec(select(Departman).where(Departman.ad == ad)).first()
    return dep.id if dep else None


def process_import_row(
    session: Session, row: dict[str, str], satir_no: int
) -> None:
    required = ("ad", "soyad", "tc_kimlik_no", "sicil_no", "rol")
    for key in required:
        if not row.get(key):
            raise ValueError(f"Eksik alan: {key}")

    try:
        rol = Rol(row["rol"].upper().replace(" ", "_"))
    except ValueError as exc:
        raise ValueError(f"Geçersiz rol: {row['rol']}") from exc

    if rol not in PERSONEL_ROLLERI:
        raise ValueError(f"Personel rolü değil: {rol.value}")

    tc = row["tc_kimlik_no"]
    sicil = row["sicil_no"]
    email = row.get("email")
    telefon = row.get("telefon")
    kullanici_adi = row.get("kullanici_adi")

    existing_sicil = session.exec(
        select(Personel).where(Personel.sicil_no == sicil)
    ).first()
    if existing_sicil:
        raise ValueError(f"Sicil zaten var: {sicil}")

    kullanici = session.exec(
        select(Kullanici).where(Kullanici.tc_kimlik_no == tc)
    ).first()

    if kullanici is not None:
        existing_personel = session.exec(
            select(Personel).where(Personel.kullanici_id == kullanici.id)
        ).first()
        if existing_personel is not None:
            raise ValueError("Bu TC için personel profili zaten var")

        # Çift profil: mevcut Hasta kullanıcısına Personel ekle
        kullanici.rol = rol
        kullanici.sifre_degistirmeli_mi = True
        kullanici.kvkk_onaylandi_mi = False
        if not kullanici.sifre_hash:
            temp = gecici_sifre_uret()
            kullanici.sifre_hash = hash_password(temp)
            _notify_credentials(kullanici, temp, email or kullanici.email, telefon)
        else:
            temp = gecici_sifre_uret()
            kullanici.sifre_hash = hash_password(temp)
            _notify_credentials(kullanici, temp, email or kullanici.email, telefon)
        if email and not kullanici.email:
            kullanici.email = email
        if telefon and not kullanici.telefon:
            kullanici.telefon = telefon
        if kullanici_adi and not kullanici.kullanici_adi:
            kullanici.kullanici_adi = kullanici_adi
        if not kullanici.ad:
            kullanici.ad = row["ad"]
        if not kullanici.soyad:
            kullanici.soyad = row["soyad"]
        session.add(kullanici)
        session.flush()
    else:
        if email:
            email_conflict = session.exec(
                select(Kullanici).where(Kullanici.email == email)
            ).first()
            if email_conflict:
                raise ValueError(f"E-posta zaten kayıtlı: {email}")
        temp = gecici_sifre_uret()
        kullanici = Kullanici(
            tc_kimlik_no=tc,
            ad=row["ad"],
            soyad=row["soyad"],
            email=email,
            telefon=telefon,
            kullanici_adi=kullanici_adi,
            sifre_hash=hash_password(temp),
            rol=rol,
            aktif_mi=True,
            sifre_degistirmeli_mi=True,
            kvkk_onaylandi_mi=False,
        )
        session.add(kullanici)
        session.flush()
        _notify_credentials(kullanici, temp, email, telefon)

    departman_id = _departman_bul(session, row.get("departman"))
    personel = Personel(
        kullanici_id=kullanici.id,
        sicil_no=sicil,
        departman_id=departman_id,
        unvan=rol.value,
    )
    session.add(personel)
    session.flush()

    if rol == Rol.DOKTOR:
        uzmanlik = row.get("uzmanlik_alani")
        diploma = row.get("diploma_no")
        if not uzmanlik or not diploma:
            raise ValueError(
                "DOKTOR için uzmanlik_alani ve diploma_no zorunludur"
            )
        diploma_exists = session.exec(
            select(Doktor).where(Doktor.diploma_no == diploma)
        ).first()
        if diploma_exists:
            raise ValueError(f"Diploma no zaten var: {diploma}")
        session.add(
            Doktor(
                personel_id=personel.id,
                uzmanlik_alani=uzmanlik,
                diploma_no=diploma,
            )
        )

    session.commit()


def _notify_credentials(
    kullanici: Kullanici,
    temp_sifre: str,
    email: str | None,
    telefon: str | None,
) -> None:
    bildirim = get_bildirim()
    mesaj = (
        f"Hastane paneli geçici şifreniz: {temp_sifre}. "
        "İlk girişte şifrenizi değiştirmeniz ve KVKK onayını tamamlamanız gerekir."
    )
    if telefon:
        bildirim.sms_gonder(telefon, mesaj)
    if email:
        bildirim.email_gonder(email, "Geçici giriş şifreniz", mesaj)


def run_import_job(session: Session, isi_id: int, rows: list[dict[str, str]]) -> None:
    isi = session.get(PersonelImportIsi, isi_id)
    if isi is None:
        return
    isi.durum = ImportDurum.ISLENIYOR
    isi.updated_at = utc_now()
    session.add(isi)
    session.commit()

    hatalar: list[dict[str, Any]] = list(isi.hata_detay or [])
    basarili = isi.basarili
    basarisiz = isi.basarisiz

    for idx, row in enumerate(rows, start=2):  # 1=header varsayımı
        try:
            process_import_row(session, row, idx)
            basarili += 1
        except Exception as exc:  # noqa: BLE001 — satır hatası, iş devam
            session.rollback()
            basarisiz += 1
            hatalar.append({"satir": idx, "hata": str(exc), "veri": row})
        isi = session.get(PersonelImportIsi, isi_id)
        if isi:
            isi.basarili = basarili
            isi.basarisiz = basarisiz
            isi.hata_detay = hatalar
            isi.updated_at = utc_now()
            session.add(isi)
            session.commit()

    isi = session.get(PersonelImportIsi, isi_id)
    if isi:
        isi.durum = ImportDurum.TAMAMLANDI
        isi.updated_at = utc_now()
        session.add(isi)
        session.commit()


async def read_upload(file: UploadFile) -> tuple[str, bytes]:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Dosya boş")
    return file.filename or "import.csv", content
